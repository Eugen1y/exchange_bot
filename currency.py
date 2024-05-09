from datetime import datetime, date
from bs4 import BeautifulSoup
import aiohttp
from db import Currency, session
from openpyxl import Workbook


def validate_exchange_rate(exchange_rate):
    """
    Validates the exchange rate.

    :param exchange_rate: The exchange rate to validate.
    :return: True if the exchange rate is valid, False otherwise.
    """

    try:
        float(exchange_rate)
        return True
    except ValueError:
        return False


async def process_html(html_text):
    """
    Processes the HTML text to extract the exchange rate.

    :param html_text: The HTML text.
    :return: The extracted exchange rate as a string, or None if not found or invalid.
    """

    soup = BeautifulSoup(html_text, 'html.parser')
    exchange_rate = soup.find('div', {'data-target': 'UAH'}).get('data-last-price')
    return str(round(float(exchange_rate), 4)) if validate_exchange_rate(exchange_rate) else None


async def make_answer():
    """
    Fetches the exchange rate from a website and saves it to the database.

    :return: A tuple containing the current datetime and exchange rate if successful, otherwise (None, None).
    """

    url = 'https://www.google.com/finance/quote/USD-UAH'

    async with aiohttp.ClientSession() as http_session:
        async with http_session.get(url) as response:
            html_text = await response.text()
            exchange_rate = await process_html(html_text)
            datetime_now = datetime.now()
            if exchange_rate:
                currency = Currency(datetime=datetime_now, exchange_rate=exchange_rate)
                session.add(currency)
                session.commit()
                return datetime_now, exchange_rate
            else:
                return None, None


async def generate_excel():
    """
    Generates an Excel file containing exchange rate data for the current day.

    :return: The filename of the generated Excel file.
    """

    wb = Workbook()
    ws = wb.active

    ws['A1'] = 'datetime'
    ws['B1'] = 'exchange_rate'

    today_date = date.today()

    rows = session.query(Currency).filter(Currency.datetime >= today_date).all()

    for idx, row in enumerate(rows, start=2):
        ws[f'A{idx}'] = row.datetime.strftime('%d.%m.%Y %H:%M:%S')
        ws[f'B{idx}'] = row.exchange_rate

    excel_filename = f'exchange_rate_{datetime.now().strftime("%Y-%m-%d_%H-%M-%S")}.xlsx'
    wb.save(excel_filename)

    return excel_filename
